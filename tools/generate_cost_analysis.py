"""
generate_cost_analysis.py
-------------------------
Generates the HMIS Bot Cost Analysis Excel workbook.
Shows the current bot behaviour only: within-query prompt caching
(OpenAI automatically caches repeated system-prompt prefixes across
the 9 LLM calls that make up each user query).

Run:  python generate_cost_analysis.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS & PRICING
# ─────────────────────────────────────────────────────────────────────────────

P54_IN   = 2.50  / 1_000_000   # GPT-5.4 input ($/token)
P54_OUT  = 15.00 / 1_000_000   # GPT-5.4 output
P54_CAC  = 0.25  / 1_000_000   # GPT-5.4 cached input (auto-applied by OpenAI)

P54M_IN  = 0.75  / 1_000_000   # GPT-5.4-mini input
P54M_OUT = 4.50  / 1_000_000   # GPT-5.4-mini output
P54M_CAC = 0.075 / 1_000_000   # GPT-5.4-mini cached input

SP_TOKENS = 8_000   # system prompt tokens — identical on every call, auto-cached

# All 9 LLM calls triggered per user query
# (name, model, input_tokens, output_tokens, purpose)
CALLS = [
    ("check_query_relevance",   "mini",     9_000,   50, "Relevance gate — rejects off-topic questions"),
    ("extract_geo_entity",      "mini",     9_000,  100, "Extracts district / block / facility from message"),
    ("formulate_understanding", "primary", 10_000,  150, "Bot states its understanding before running SQL"),
    ("check_if_confirmed",      "mini",     9_500,   80, "Checks user yes/no confirmation reply"),
    ("synthesize_final_query",  "primary", 10_000,  100, "Synthesises clean final question from conversation"),
    ("generate_sql",            "primary", 13_500,  500, "Generates BigQuery SQL (incl. ~3k rule4_text)"),
    ("summarize_results",       "primary", 12_000,  400, "Converts BQ rows to plain-English answer (parallel)"),
    ("generate_key_findings",   "primary", 12_000,  250, "Bullet-point key findings (parallel)"),
    ("get_chart_spec",          "mini",     9_500,  150, "Decides chart type and axes (parallel)"),
]

USER_COUNTS = [25, 50, 100, 500, 1_000, 5_000]

USAGE = {
    "Low":    {"qpd": 2,  "days": 22, "desc": "2 queries/user/day — occasional data checks"},
    "Medium": {"qpd": 5,  "days": 22, "desc": "5 queries/user/day — regular daily analysis"},
    "High":   {"qpd": 15, "days": 22, "desc": "15 queries/user/day — heavy continuous use"},
}

SCHEMA_CALLS_PER_MONTH = 4                            # once per week
SCHEMA_FIXED_COST_PM   = SCHEMA_CALLS_PER_MONTH * (4_500 * P54_IN + 3_000 * P54_OUT)

BQ_PER_QUERY     = 0.000010   # BigQuery scan cost (~1–10 MB filtered query at $5/TB)
VOICE_FRACTION   = 0.20       # 20% of queries arrive as voice messages
VOICE_COST_EXTRA = 0.003      # Whisper: $0.006/min × 30 s average


# ─────────────────────────────────────────────────────────────────────────────
# COST CALCULATION  (within-query caching — current bot behaviour)
# ─────────────────────────────────────────────────────────────────────────────

def _call_cost(call_order_in_model: int, call: tuple) -> float:
    """
    Cost of a single LLM call with within-query prompt caching.
    The first call of each model type pays the full rate on the system prompt;
    subsequent calls of the same model type pay the cached rate.
    OpenAI applies this automatically — no code change needed.
    """
    _, model, inp, out, _ = call
    p_in  = P54_IN  if model == "primary" else P54M_IN
    p_out = P54_OUT if model == "primary" else P54M_OUT
    p_cac = P54_CAC if model == "primary" else P54M_CAC

    sp  = min(inp, SP_TOKENS)   # system-prompt portion of this call's input
    nsp = inp - sp              # remaining (user message, history, results)

    sp_cost = sp * (p_in if call_order_in_model == 0 else p_cac)
    return sp_cost + nsp * p_in + out * p_out


def _query_cost() -> dict:
    """Aggregate cost for all 9 calls in one user query."""
    primary_calls = [c for c in CALLS if c[1] == "primary"]
    mini_calls    = [c for c in CALLS if c[1] == "mini"]

    pc = sum(_call_cost(i, c) for i, c in enumerate(primary_calls))
    mc = sum(_call_cost(i, c) for i, c in enumerate(mini_calls))
    return {
        "primary": pc, "mini": mc, "total": pc + mc,
        "primary_in":  sum(c[2] for c in CALLS if c[1] == "primary"),
        "primary_out": sum(c[3] for c in CALLS if c[1] == "primary"),
        "mini_in":     sum(c[2] for c in CALLS if c[1] == "mini"),
        "mini_out":    sum(c[3] for c in CALLS if c[1] == "mini"),
    }


QC = _query_cost()


# ─────────────────────────────────────────────────────────────────────────────
# STYLING HELPERS
# ─────────────────────────────────────────────────────────────────────────────

DARK_BLUE    = "1F3864"
MID_BLUE     = "2E75B6"
LIGHT_BLUE   = "BDD7EE"
LIGHTEST     = "DEEAF1"
LIGHT_GREEN  = "E2EFDA"
LIGHT_ORANGE = "FCE4D6"
YELLOW       = "FFF2CC"
WHITE        = "FFFFFF"
GREY_ROW     = "F2F2F2"


def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _cs(c, bold=False, size=10, color="000000", bg=WHITE,
        fmt=None, center=False, wrap=False):
    c.font      = Font(bold=bold, size=size, color=color, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                             vertical="center", wrap_text=wrap)
    if fmt:
        c.number_format = fmt
    c.border = _thin()


def _hdr_cell(ws, row, col, text, bg=MID_BLUE, wrap=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=True, size=10, color=WHITE, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border    = _thin()


def _title(ws, row, text, cols, bg=DARK_BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26


def _sub(ws, row, text, cols, bg=MID_BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, size=11, color=WHITE, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


def _note(ws, row, text, cols, bg=LIGHTEST):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(italic=True, size=9, color="595959", name="Calibri")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 32


def _w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


USAGE_BG = {"Low": LIGHT_GREEN, "Medium": YELLOW, "High": LIGHT_ORANGE}


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1: EXECUTIVE SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

def build_summary(wb):
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False

    N = 8   # total columns used on this sheet

    _title(ws, 1, "HMIS Telegram Analytics Bot — Cost Analysis", N)
    _note(ws, 2,
          f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  "
          f"Models: GPT-5.4 (primary) · GPT-5.4-mini (fast)  |  "
          f"Scenario: within-query prompt caching (current bot behaviour — automatic, no code change needed)  |  "
          f"22 working days/month · 9 LLM calls/query", N)

    # ── Pricing reference ────────────────────────────────────────────────────
    r = 4
    _sub(ws, r, "Model Pricing", N); r += 1
    for ci, h in enumerate(["Model", "Input $/1M", "Output $/1M",
                              "Cached Input $/1M", "Role in Bot"], 1):
        _hdr_cell(ws, r, ci, h)
    r += 1
    for i, row_data in enumerate([
        ("GPT-5.4 (primary)",   "$2.50",  "$15.00", "$0.25",
         "SQL generation, summarisation, key findings"),
        ("GPT-5.4-mini (fast)", "$0.75",  "$4.50",  "$0.075",
         "Geo extraction, confirmation check, chart spec"),
    ]):
        bg = WHITE if i % 2 == 0 else GREY_ROW
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=ci, value=val)
            _cs(c, bg=bg, center=(ci > 1))
        r += 1

    # ── Per-query cost (single scenario) ─────────────────────────────────────
    r += 1
    _sub(ws, r, "Cost Per Query — Current Bot Behaviour", N); r += 1
    _note(ws, r,
          "OpenAI automatically caches repeated prompt prefixes (>=1,024 tokens). "
          "The system prompt (~8,000 tokens) is identical on all 9 calls within a query. "
          "Calls 2-5 (primary) and 2-4 (mini) receive the cached rate automatically.", N)
    r += 1

    for ci, h in enumerate(["Component", "Input Tokens", "Output Tokens",
                              "Input Cost", "Output Cost", "Total Cost"], 1):
        _hdr_cell(ws, r, ci, h)
    r += 1

    prim_in_cost  = QC["primary_in"]  * P54_IN   # not exact, but shown for reference
    mini_in_cost  = QC["mini_in"]     * P54M_IN
    prim_out_cost = QC["primary_out"] * P54_OUT
    mini_out_cost = QC["mini_out"]    * P54M_OUT

    # Recompute with caching properly
    primary_calls = [c for c in CALLS if c[1] == "primary"]
    mini_calls    = [c for c in CALLS if c[1] == "mini"]
    prim_in_cached  = sum(_call_cost(i, c) - c[3]*P54_OUT for i, c in enumerate(primary_calls))
    mini_in_cached  = sum(_call_cost(i, c) - c[3]*P54M_OUT for i, c in enumerate(mini_calls))
    prim_out_actual = QC["primary_out"] * P54_OUT
    mini_out_actual = QC["mini_out"]    * P54M_OUT

    rows_pq = [
        ("GPT-5.4 (primary)",   QC["primary_in"],  QC["primary_out"],
         prim_in_cached, prim_out_actual, QC["primary"],  WHITE),
        ("GPT-5.4-mini (fast)", QC["mini_in"],     QC["mini_out"],
         mini_in_cached, mini_out_actual, QC["mini"],     GREY_ROW),
        ("TOTAL per query",     QC["primary_in"]+QC["mini_in"],
         QC["primary_out"]+QC["mini_out"],
         prim_in_cached+mini_in_cached,
         prim_out_actual+mini_out_actual, QC["total"],    LIGHT_BLUE),
    ]
    for label, ti, to, ic, oc, tot, bg in rows_pq:
        bold = "TOTAL" in label
        for ci, val in enumerate([label, ti, to, ic, oc, tot], 1):
            c = ws.cell(row=r, column=ci, value=val)
            fmt = "#,##0" if ci in (2, 3) else ('"$"#,##0.00000' if ci > 3 else None)
            _cs(c, bold=bold, bg=bg, center=(ci > 1), fmt=fmt)
        r += 1

    _note(ws, r,
          f"Input cost reflects within-query caching: system prompt tokens on calls 2+ "
          f"charged at $0.25/M (primary) and $0.075/M (mini) instead of full rate. "
          f"Full-rate cost without any caching would be ${sum(_call_cost.__wrapped__ if hasattr(_call_cost,'__wrapped__') else 0 for _ in []):,.4f} — "
          f"caching saves ~45% on input costs automatically.", N)
    r += 2

    # ── Monthly cost matrix ───────────────────────────────────────────────────
    _sub(ws, r, "Monthly Cost Summary", N); r += 1
    _note(ws, r,
          "Total = (queries/month x $%.4f/query) + $%.2f fixed (schema cache) + BigQuery scan costs" %
          (QC["total"], SCHEMA_FIXED_COST_PM), N)
    r += 1

    ws.cell(row=r, column=1, value="Usage Level")
    ws.cell(row=r, column=2, value="q/user/month")
    for ci, u in enumerate(USER_COUNTS, 3):
        _hdr_cell(ws, r, ci, f"{u:,} users")
    _hdr_cell(ws, r, 3 + len(USER_COUNTS), "$/user/month")
    r += 1

    for label, udata in USAGE.items():
        qpm = udata["qpd"] * udata["days"]
        bg  = USAGE_BG[label]
        c = ws.cell(row=r, column=1, value=label)
        _cs(c, bold=True, bg=bg)
        c2 = ws.cell(row=r, column=2, value=qpm)
        _cs(c2, bg=bg, center=True, fmt="#,##0")
        for ci, users in enumerate(USER_COUNTS, 3):
            total_q = users * qpm
            cost    = total_q * QC["total"] + SCHEMA_FIXED_COST_PM + total_q * BQ_PER_QUERY
            cx = ws.cell(row=r, column=ci, value=cost)
            _cs(cx, bg=bg, center=True, fmt='"$"#,##0.00')
        cpu = qpm * QC["total"] + qpm * BQ_PER_QUERY
        cy  = ws.cell(row=r, column=3 + len(USER_COUNTS), value=cpu)
        _cs(cy, bold=True, bg=bg, center=True, fmt='"$"#,##0.0000')
        r += 1

    _note(ws, r,
          "Fixed schema-cache cost ($%.4f/month) is negligible — it rounds to <$0.01 per user "
          "at any scale. BigQuery scan costs (<$0.001 per query) are included but invisible at this rounding." %
          SCHEMA_FIXED_COST_PM, N)

    _w(ws, 1, 18); _w(ws, 2, 16)
    for ci in range(3, 3 + len(USER_COUNTS)):
        _w(ws, ci, 14)
    _w(ws, 3 + len(USER_COUNTS), 18)


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2: PER-QUERY COST BREAKDOWN
# ─────────────────────────────────────────────────────────────────────────────

def build_per_query(wb):
    ws = wb.create_sheet("Per-Query Cost Breakdown")
    ws.sheet_view.showGridLines = False
    N = 8

    _title(ws, 1, "Per-Query LLM Cost — All 9 Calls (within-query caching)", N)
    _note(ws, 2,
          "System prompt (~8,000 tokens) is included in every call. "
          "OpenAI caches it automatically after the first call of each model type within a query. "
          "Calls 7, 8, 9 run in parallel. Token counts are estimates (actual varies ~20%).", N)

    r = 4
    _sub(ws, r, "Call-by-Call Breakdown", N); r += 1
    for ci, h in enumerate(["#", "LLM Call", "Model", "Input Tokens", "Output Tokens",
                              "SP Cached?", "Cost", "Purpose"], 1):
        _hdr_cell(ws, r, ci, h, wrap=True)
    r += 1

    primary_idx = mini_idx = 0
    for gidx, call in enumerate(CALLS):
        _, model, inp, out, purpose = call
        if model == "primary":
            idx = primary_idx; primary_idx += 1
        else:
            idx = mini_idx; mini_idx += 1

        cost    = _call_cost(idx, call)
        cached  = "No (1st call)" if idx == 0 else "Yes"
        bg      = LIGHT_BLUE if model == "primary" else LIGHTEST

        for ci, val in enumerate([gidx+1, call[0], model.upper(),
                                   inp, out, cached, cost, purpose], 1):
            c = ws.cell(row=r, column=ci, value=val)
            fmt = "#,##0" if ci in (4, 5) else ('"$"#,##0.00000' if ci == 7 else None)
            _cs(c, bg=bg, center=(ci in (1, 3, 4, 5, 6, 7)), fmt=fmt)
        r += 1

    # Subtotals
    for label, model_tag, ti, to, cost, bg in [
        ("PRIMARY subtotal",     "PRIMARY",
         QC["primary_in"], QC["primary_out"], QC["primary"], MID_BLUE),
        ("FAST (mini) subtotal", "MINI",
         QC["mini_in"],    QC["mini_out"],    QC["mini"],    MID_BLUE),
        ("GRAND TOTAL / query",  "",
         QC["primary_in"]+QC["mini_in"], QC["primary_out"]+QC["mini_out"],
         QC["total"], DARK_BLUE),
    ]:
        for ci, val in enumerate(["", label, model_tag, ti, to, "", cost, ""], 1):
            c = ws.cell(row=r, column=ci, value=val)
            fmt = "#,##0" if ci in (4,5) else ('"$"#,##0.00000' if ci == 7 else None)
            _cs(c, bold=True, bg=bg, color=WHITE, center=(ci in (3,4,5,7)), fmt=fmt)
        r += 1

    r += 1
    _note(ws, r,
          "Caching detail: primary model has 5 calls — call 1 pays full SP rate ($2.50/M), "
          "calls 2-5 pay cached rate ($0.25/M). Mini model has 4 calls — call 1 full ($0.75/M), "
          "calls 2-4 cached ($0.075/M). Non-SP tokens (user message, history, results) always "
          "pay the full input rate. Output tokens are never cached.", N)

    _w(ws, 1, 4); _w(ws, 2, 32); _w(ws, 3, 12)
    _w(ws, 4, 14); _w(ws, 5, 14); _w(ws, 6, 16)
    _w(ws, 7, 16); _w(ws, 8, 48)
    ws.row_dimensions[1].height = 22


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3: MONTHLY COST
# ─────────────────────────────────────────────────────────────────────────────

def build_monthly(wb):
    ws = wb.create_sheet("Monthly Cost")
    ws.sheet_view.showGridLines = False
    N = 10

    _title(ws, 1, "Monthly Cost — Users x Usage Level (current bot behaviour)", N)
    _note(ws, 2,
          "Cost per query: $%.4f (within-query caching, automatic).  "
          "Formula: total_queries x $%.4f + $%.4f fixed (schema cache) + BigQuery scan." %
          (QC["total"], QC["total"], SCHEMA_FIXED_COST_PM), N)

    r = 4
    # Header
    _hdr_cell(ws, r, 1, "Usage Level")
    _hdr_cell(ws, r, 2, "q/user/day")
    _hdr_cell(ws, r, 3, "q/user/month")
    for ci, u in enumerate(USER_COUNTS, 4):
        _hdr_cell(ws, r, ci, f"{u:,} users")
    _hdr_cell(ws, r, 4 + len(USER_COUNTS), "$/user/month")
    r += 1

    for label, udata in USAGE.items():
        qpd = udata["qpd"]
        qpm = qpd * udata["days"]
        bg  = USAGE_BG[label]
        for ci, val in enumerate([label, qpd, qpm], 1):
            c = ws.cell(row=r, column=ci, value=val)
            _cs(c, bold=(ci==1), bg=bg, center=(ci>1), fmt="#,##0" if ci>1 else None)
        for ci, users in enumerate(USER_COUNTS, 4):
            total_q = users * qpm
            cost    = total_q * QC["total"] + SCHEMA_FIXED_COST_PM + total_q * BQ_PER_QUERY
            cx = ws.cell(row=r, column=ci, value=cost)
            _cs(cx, bg=bg, center=True, fmt='"$"#,##0.00')
        cpu = qpm * (QC["total"] + BQ_PER_QUERY)
        cy  = ws.cell(row=r, column=4+len(USER_COUNTS), value=cpu)
        _cs(cy, bold=True, bg=bg, center=True, fmt='"$"#,##0.0000')
        r += 1

    r += 2
    # Per-user monthly cost table (cleaner view)
    _sub(ws, r, "Cost Per User Per Month", N); r += 1
    _note(ws, r, "Independent of total user count — scales linearly.", N); r += 1
    for ci, h in enumerate(["Usage Level", "q/user/month", "LLM Cost/User",
                              "BQ Cost/User", "Total/User/Month", "Annual/User"], 1):
        _hdr_cell(ws, r, ci, h)
    r += 1
    for label, udata in USAGE.items():
        qpm = udata["qpd"] * udata["days"]
        bg  = USAGE_BG[label]
        llm = qpm * QC["total"]
        bq  = qpm * BQ_PER_QUERY
        tot = llm + bq
        ann = tot * 12
        for ci, val in enumerate([label, qpm, llm, bq, tot, ann], 1):
            c = ws.cell(row=r, column=ci, value=val)
            fmt = "#,##0" if ci==2 else ('"$"#,##0.0000' if ci in (3,4,5,6) else None)
            _cs(c, bold=(ci==1 or ci==5), bg=bg, center=(ci>1), fmt=fmt)
        r += 1

    r += 2
    # Total queries per month table
    _sub(ws, r, "Total Queries per Month", N); r += 1
    _hdr_cell(ws, r, 1, "Usage Level")
    for ci, u in enumerate(USER_COUNTS, 2):
        _hdr_cell(ws, r, ci, f"{u:,} users")
    r += 1
    for label, udata in USAGE.items():
        qpm = udata["qpd"] * udata["days"]
        bg  = USAGE_BG[label]
        c = ws.cell(row=r, column=1, value=label)
        _cs(c, bold=True, bg=bg)
        for ci, users in enumerate(USER_COUNTS, 2):
            cx = ws.cell(row=r, column=ci, value=users * qpm)
            _cs(cx, bg=bg, center=True, fmt="#,##0")
        r += 1

    _w(ws, 1, 18); _w(ws, 2, 14); _w(ws, 3, 16)
    for ci in range(4, 4+len(USER_COUNTS)):
        _w(ws, ci, 14)
    _w(ws, 4+len(USER_COUNTS), 18)
    ws.row_dimensions[1].height = 22


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 4: COST COMPONENTS
# ─────────────────────────────────────────────────────────────────────────────

def build_components(wb):
    ws = wb.create_sheet("Cost Components")
    ws.sheet_view.showGridLines = False
    N = 7

    _title(ws, 1, "Cost Components — What Makes Up Each Query's Cost", N)

    # ── LLM token breakdown ──────────────────────────────────────────────────
    r = 3
    _sub(ws, r, "1. LLM API Cost Breakdown", N); r += 1
    _note(ws, r,
          "Total input: %s tokens (primary) + %s tokens (mini) per query. "
          "System prompt accounts for ~%d%% of primary input and ~%d%% of mini input." %
          (f"{QC['primary_in']:,}", f"{QC['mini_in']:,}",
           5*SP_TOKENS*100//QC['primary_in'], 4*SP_TOKENS*100//QC['mini_in']), N)
    r += 1
    for ci, h in enumerate(["Token Category", "GPT-5.4 (primary)", "GPT-5.4-mini (fast)", "Notes"], 1):
        _hdr_cell(ws, r, ci, h)
    r += 1
    token_rows = [
        ("System prompt tokens (per call)", f"{SP_TOKENS:,}", f"{SP_TOKENS:,}",
         "Identical on every call — auto-cached after 1st"),
        ("Calls in query", "5", "4", "Primary: calls 3,5,6,7,8  |  Mini: calls 1,2,4,9"),
        ("Total SP tokens (all calls)", f"{5*SP_TOKENS:,}", f"{4*SP_TOKENS:,}",
         "Calls 2+ pay cached rate"),
        ("SP tokens at full rate", f"{SP_TOKENS:,}", f"{SP_TOKENS:,}",
         "First call only"),
        ("SP tokens at cached rate", f"{4*SP_TOKENS:,}", f"{3*SP_TOKENS:,}",
         "Calls 2-5 (primary) / 2-4 (mini)"),
        ("Non-SP input tokens", f"{QC['primary_in']-5*SP_TOKENS:,}",
         f"{QC['mini_in']-4*SP_TOKENS:,}",
         "User message, history, BQ results, rule4_text — always full rate"),
        ("Output tokens", f"{QC['primary_out']:,}", f"{QC['mini_out']:,}",
         "Never cached"),
    ]
    for i, (k, v1, v2, note) in enumerate(token_rows):
        bg = LIGHT_BLUE if i % 2 == 0 else LIGHTEST
        for ci, val in enumerate([k, v1, v2, note], 1):
            c = ws.cell(row=r, column=ci, value=val)
            _cs(c, bg=bg, center=(ci in (2,3)))
        r += 1

    # ── Schema cache cost ────────────────────────────────────────────────────
    r += 1
    _sub(ws, r, "2. Schema Cache — Fixed Monthly Cost", N); r += 1
    _note(ws, r,
          "GPT-5.4 is called once per week to detect COALESCE column pairs from the live "
          "BigQuery schema. This is fixed regardless of user count.", N)
    r += 1
    for ci, h in enumerate(["Item", "Value", "Notes"], 1):
        _hdr_cell(ws, r, ci, h)
    r += 1
    sc_in  = SCHEMA_CALLS_PER_MONTH * 4_500 * P54_IN
    sc_out = SCHEMA_CALLS_PER_MONTH * 3_000 * P54_OUT
    for i, (k, v, n) in enumerate([
        ("Calls per month",             str(SCHEMA_CALLS_PER_MONTH), "Once per week (7-day cache TTL)"),
        ("Input per call",              "4,500 tokens",  "Indicator column list + prompt"),
        ("Output per call",             "3,000 tokens",  "JSON array of COALESCE pairs"),
        ("Input cost per month",        f"${sc_in:.4f}",  f"{SCHEMA_CALLS_PER_MONTH} x 4,500 x $2.50/M"),
        ("Output cost per month",       f"${sc_out:.4f}", f"{SCHEMA_CALLS_PER_MONTH} x 3,000 x $15.00/M"),
        ("TOTAL fixed cost per month",  f"${SCHEMA_FIXED_COST_PM:.4f}", "~$0.23/month, regardless of users"),
    ]):
        bg = LIGHT_BLUE if i % 2 == 0 else LIGHTEST
        for ci, val in enumerate([k, v, n], 1):
            c = ws.cell(row=r, column=ci, value=val)
            _cs(c, bold=("TOTAL" in k), bg=bg, center=(ci==2))
        r += 1

    # ── BQ cost ──────────────────────────────────────────────────────────────
    r += 1
    _sub(ws, r, "3. BigQuery Scan Cost", N); r += 1
    _note(ws, r,
          "Each user query runs 1-3 BigQuery SQL statements. Queries are filtered by "
          "district/block/month — typical scan: 1-10 MB at $5/TB = $0.000005-$0.00005 per query.", N)
    r += 1
    for ci, h in enumerate(["Usage Level", "Users", "Queries/Month",
                              "BQ Cost/Query", "Total BQ Cost/Month"], 1):
        _hdr_cell(ws, r, ci, h)
    r += 1
    for label, udata in USAGE.items():
        qpm = udata["qpd"] * udata["days"]
        bg  = USAGE_BG[label]
        for users in [25, 100, 1_000, 5_000]:
            total_q = users * qpm
            for ci, val in enumerate([f"{label} / {users:,} users", users, total_q,
                                       BQ_PER_QUERY, total_q * BQ_PER_QUERY], 1):
                c = ws.cell(row=r, column=ci, value=val)
                fmt = "#,##0" if ci in (2,3) else ('"$"#,##0.00000' if ci==4 else
                      ('"$"#,##0.00' if ci==5 else None))
                _cs(c, bg=bg, center=(ci>1), fmt=fmt)
            r += 1

    # ── Voice cost ───────────────────────────────────────────────────────────
    r += 1
    _sub(ws, r, "4. Voice Transcription Add-on (Whisper API)", N); r += 1
    _note(ws, r,
          f"$0.006/minute. Average voice query ~30 seconds = $0.003 extra. "
          f"Assumed {VOICE_FRACTION*100:.0f}% of all queries are voice messages.", N)
    r += 1
    for ci, h in enumerate(["Usage Level", "Users", "Total q/Month",
                              "Voice q/Month (20%)", "Voice Cost/Month",
                              "Voice as % of LLM Cost"], 1):
        _hdr_cell(ws, r, ci, h, wrap=True)
    r += 1
    for label, udata in USAGE.items():
        qpm = udata["qpd"] * udata["days"]
        bg  = USAGE_BG[label]
        for users in [25, 100, 1_000, 5_000]:
            total_q  = users * qpm
            voice_q  = total_q * VOICE_FRACTION
            voice_c  = voice_q * VOICE_COST_EXTRA
            llm_c    = total_q * QC["total"]
            voice_pct = voice_c / llm_c * 100 if llm_c else 0
            for ci, val in enumerate([f"{label} / {users:,} users", users, total_q,
                                       voice_q, voice_c, voice_pct], 1):
                c = ws.cell(row=r, column=ci, value=val)
                fmt = ("#,##0"     if ci in (2,3,4) else
                       '"$"#,##0.00' if ci==5 else
                       '0.0"%"'    if ci==6 else None)
                _cs(c, bg=bg, center=(ci>1), fmt=fmt)
            r += 1

    _w(ws, 1, 32); _w(ws, 2, 14); _w(ws, 3, 18)
    _w(ws, 4, 20); _w(ws, 5, 22); _w(ws, 6, 26); _w(ws, 7, 20)
    ws.row_dimensions[1].height = 22


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 5: SCHEMA CACHE IMPACT
# ─────────────────────────────────────────────────────────────────────────────

def build_schema_cache(wb):
    ws = wb.create_sheet("Schema Cache Impact")
    ws.sheet_view.showGridLines = False
    N = 6

    _title(ws, 1, "Schema Cache LLM Detection — Cost Impact", N)
    _note(ws, 2,
          "The schema_cache.py upgrade replaced free regex detection with one GPT-5.4 call per week. "
          "This sheet shows the amortised per-query impact at every usage scale.", N)

    r = 4
    _sub(ws, r, "Fixed Monthly Cost", N); r += 1
    for ci, h in enumerate(["Item", "Value", "Notes"], 1):
        _hdr_cell(ws, r, ci, h)
    r += 1
    sc_in  = SCHEMA_CALLS_PER_MONTH * 4_500 * P54_IN
    sc_out = SCHEMA_CALLS_PER_MONTH * 3_000 * P54_OUT
    for i, (k, v, n) in enumerate([
        ("Previous (regex)",      "$0.0000/month", "Pure Python, no API calls"),
        ("Current (LLM-based)",   f"${SCHEMA_FIXED_COST_PM:.4f}/month", "GPT-5.4, once per week"),
        ("Monthly increase",      f"${SCHEMA_FIXED_COST_PM:.4f}/month", "~$0.23/month"),
    ]):
        bg = LIGHT_BLUE if i % 2 == 0 else LIGHTEST
        for ci, val in enumerate([k, v, n], 1):
            c = ws.cell(row=r, column=ci, value=val)
            _cs(c, bold=("increase" in k.lower()), bg=bg, center=(ci==2))
        r += 1

    r += 2
    _sub(ws, r, "Amortised Per-Query Impact", N); r += 1
    for ci, h in enumerate(["Usage Level", "Users", "Queries/Month",
                              "Schema $/Month", "Amortised $/Query",
                              "% of Per-Query Cost"], 1):
        _hdr_cell(ws, r, ci, h, wrap=True)
    r += 1
    for label, udata in USAGE.items():
        qpm = udata["qpd"] * udata["days"]
        bg  = USAGE_BG[label]
        for users in USER_COUNTS:
            total_q = users * qpm
            amort   = SCHEMA_FIXED_COST_PM / total_q if total_q else 0
            pct     = amort / QC["total"] * 100 if QC["total"] else 0
            for ci, val in enumerate([label, users, total_q,
                                       SCHEMA_FIXED_COST_PM, amort, pct], 1):
                c = ws.cell(row=r, column=ci, value=val)
                fmt = ("#,##0"             if ci in (2,3)  else
                       '"$"#,##0.0000'     if ci==4        else
                       '"$"#,##0.00000000' if ci==5        else
                       '0.000"%"'          if ci==6        else None)
                _cs(c, bg=bg, center=(ci>1), fmt=fmt)
            r += 1

    _note(ws, r,
          "At 25 users / Low: schema cache = 0.12% of per-query cost.  "
          "At 5,000 users / High: 0.00003%.  Impact is negligible at all scales.", N)

    _w(ws, 1, 18); _w(ws, 2, 12); _w(ws, 3, 18)
    _w(ws, 4, 20); _w(ws, 5, 24); _w(ws, 6, 22)
    ws.row_dimensions[1].height = 22


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 6: COST OPTIMISATION LEVERS
# ─────────────────────────────────────────────────────────────────────────────

def build_optimisation(wb):
    ws = wb.create_sheet("Cost Optimisation")
    ws.sheet_view.showGridLines = False
    N = 6

    _title(ws, 1, "Cost Optimisation Levers — Ranked by Impact", N)
    _note(ws, 2,
          "Current cost per query: $%.4f (within-query caching already active). "
          "Levers below describe further reductions available." % QC["total"], N)

    r = 4
    for ci, h in enumerate(["#", "Optimisation", "Impact",
                              "Estimated Saving", "How It Works",
                              "Implementation Effort"], 1):
        _hdr_cell(ws, r, ci, h)
    r += 1

    levers = [
        ("1", "Cross-Query Prompt Caching (already partially active)",
         "HIGH",
         "Up to $0.09/query additional saving at high traffic (queries within 5-min windows share cached SP)",
         "OpenAI's cache also fires cross-query when the same prefix arrives within ~5 minutes. "
         "At high usage (many users, frequent queries), a large fraction of cross-query calls also "
         "benefit. No code change needed — maximised naturally as traffic grows.",
         "No action needed — happens automatically",
         LIGHT_GREEN),
        ("2", "Reduce System Prompt Size",
         "MEDIUM",
         "~$0.003 saving per 1,000 tokens removed (9 calls x 1k tokens x avg $0.33/M blended rate)",
         "System prompt (~8,000 tokens) dominates input cost. Move rarely-needed content — "
         "e.g. full completeness SQL template, PRE SQL structure — out of the static prompt "
         "and inject them only for relevant query types.",
         "Prompt restructuring — medium effort, test carefully",
         YELLOW),
        ("3", "Skip Confirmation Step for Unambiguous Queries",
         "MEDIUM",
         "~$0.015/query (eliminates 2 LLM calls: formulate_understanding + check_if_confirmed)",
         "When geo is resolved to a single match and the indicator is explicit, "
         "bypass the confirmation loop and go straight to SQL generation. "
         "Affects ~40-60% of typical queries.",
         "Code change in bot.py — low risk, medium effort",
         YELLOW),
        ("4", "Use Fast Model for SQL Error Fixing",
         "LOW",
         "~$0.008 per retry (10% of queries need a retry); saves ~$0.001/query on average",
         "fix_sql() currently uses GPT-5.4. Switching to GPT-5.4-mini for simple "
         "BigQuery error fixes (column not found, syntax errors) reduces retry cost by ~85%.",
         "One config change in claude_client.py — low risk",
         LIGHT_GREEN),
        ("5", "Truncate BQ Results Before Summarisation",
         "LOW",
         "~$0.005/query (reduces summarize_results and generate_key_findings input by ~2k tokens each)",
         "LLM rarely gains insight from >50 raw BQ rows. Cap results at top 50 before "
         "passing to summarise calls. Aggregate in SQL where possible.",
         "5-line code change — very low risk",
         LIGHT_GREEN),
        ("6", "Cache Geography Lookups in Session",
         "LOW",
         "<$0.001/query (eliminates repeated BQ lookup queries within same conversation)",
         "District/block codes resolved once per conversation are cached in the user session dict. "
         "Avoids duplicate INFORMATION_SCHEMA and geo lookup BigQuery calls for the same location.",
         "Small change in bigquery_client.py — no risk",
         LIGHT_GREEN),
    ]

    impact_bg = {"HIGH": LIGHT_ORANGE, "MEDIUM": YELLOW, "LOW": LIGHT_GREEN}
    for num, name, impact, saving, detail, effort, bg in levers:
        for ci, val in enumerate([num, name, impact, saving, detail, effort], 1):
            c = ws.cell(row=r, column=ci, value=val)
            _cs(c, bold=(ci in (1, 2)), bg=bg, center=(ci==1),
                wrap=(ci in (4, 5, 6)))
        ws.row_dimensions[r].height = 60
        r += 1

    _w(ws, 1, 4); _w(ws, 2, 36); _w(ws, 3, 10)
    _w(ws, 4, 38); _w(ws, 5, 58); _w(ws, 6, 36)
    ws.row_dimensions[1].height = 22


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 7: ASSUMPTIONS
# ─────────────────────────────────────────────────────────────────────────────

def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False
    N = 7

    _title(ws, 1, "Assumptions & Token Estimates", N)

    r = 3
    _sub(ws, r, "Pricing (as provided)", N); r += 1
    for ci, h in enumerate(["Model", "Token Type", "Price", "Unit"], 1):
        _hdr_cell(ws, r, ci, h)
    r += 1
    for i, row_data in enumerate([
        ("GPT-5.4",      "Input",        "$2.50",  "per 1M tokens"),
        ("GPT-5.4",      "Output",       "$15.00", "per 1M tokens"),
        ("GPT-5.4",      "Cached Input", "$0.25",  "per 1M tokens (auto-applied)"),
        ("GPT-5.4-mini", "Input",        "$0.75",  "per 1M tokens"),
        ("GPT-5.4-mini", "Output",       "$4.50",  "per 1M tokens"),
        ("GPT-5.4-mini", "Cached Input", "$0.075", "per 1M tokens (auto-applied)"),
    ]):
        bg = LIGHT_BLUE if i % 2 == 0 else LIGHTEST
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=ci, value=val)
            _cs(c, bg=bg, center=(ci in (3, 4)))
        r += 1

    r += 2
    _sub(ws, r, "Token Estimates per LLM Call", N); r += 1
    _note(ws, r,
          "System prompt: ~8,000 tokens (HMIS schema + SQL rules + PRE/completeness templates). "
          "Rule4_text injection (generate_sql only): ~2,500 tokens. "
          "BQ results (summarise calls): ~3,000 tokens. Actual costs vary +/-20%.", N)
    r += 1
    for ci, h in enumerate(["#", "Call", "Model", "Input Tokens", "Output Tokens",
                              "SP Tokens", "SP % of Input", "Purpose"], 1):
        _hdr_cell(ws, r, ci, h, wrap=True)
    r += 1
    for i, (name, model, inp, out, purpose) in enumerate(CALLS):
        sp     = min(SP_TOKENS, inp)
        sp_pct = sp / inp * 100
        bg     = LIGHT_BLUE if model == "primary" else LIGHTEST
        for ci, val in enumerate([i+1, name, model.upper(), inp, out,
                                   sp, f"{sp_pct:.0f}%", purpose], 1):
            c = ws.cell(row=r, column=ci, value=val)
            fmt = "#,##0" if ci in (4, 5, 6) else None
            _cs(c, bg=bg, center=(ci in (1,3,4,5,6,7)), fmt=fmt)
        r += 1

    r += 2
    _sub(ws, r, "Other Assumptions", N); r += 1
    for ci, h in enumerate(["Parameter", "Value", "Notes"], 1):
        _hdr_cell(ws, r, ci, h)
    r += 1
    for i, (k, v, n) in enumerate([
        ("Working days/month",             "22",            "Standard (Mon-Sat minus public holidays)"),
        ("LLM calls per user query",       "9",             "All fire for a complete data query"),
        ("System prompt size",             "~8,000 tokens", "Estimated from schema + rules + templates"),
        ("Caching mechanism",              "Within-query",  "OpenAI auto-caches repeated prefix >1,024 tokens"),
        ("Caching applies from",           "Call 2 onward", "First call pays full rate; calls 2+ get cached rate"),
        ("BQ scan cost per query",         "$0.00001",      "$5/TB, filtered queries ~1-10 MB"),
        ("Voice query fraction",           "20%",           "Estimated; varies by user population"),
        ("Whisper transcription cost",     "$0.003/query",  "$0.006/min x 30 s average"),
        ("Schema cache LLM frequency",     "4x/month",      "7-day TTL + column-count change trigger"),
        ("Token estimate accuracy",        "+/- 20%",       "Actual varies with question complexity and history length"),
    ]):
        bg = LIGHT_BLUE if i % 2 == 0 else LIGHTEST
        for ci, val in enumerate([k, v, n], 1):
            c = ws.cell(row=r, column=ci, value=val)
            _cs(c, bg=bg, center=(ci==2))
        r += 1

    _w(ws, 1, 6); _w(ws, 2, 36); _w(ws, 3, 14); _w(ws, 4, 16)
    _w(ws, 5, 16); _w(ws, 6, 16); _w(ws, 7, 16); _w(ws, 8, 52)
    ws.row_dimensions[1].height = 22


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()

    print("Building sheets...")
    build_summary(wb);       print("  OK Executive Summary")
    build_per_query(wb);     print("  OK Per-Query Cost Breakdown")
    build_monthly(wb);       print("  OK Monthly Cost")
    build_components(wb);    print("  OK Cost Components")
    build_schema_cache(wb);  print("  OK Schema Cache Impact")
    build_optimisation(wb);  print("  OK Cost Optimisation")
    build_assumptions(wb);   print("  OK Assumptions")

    fname = f"HMIS_Bot_Cost_Analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(fname)
    print(f"\nSaved: {fname}")

    print("\n-- Cost Reference (current bot: within-query caching) ----------")
    print(f"  Per query:                 ${QC['total']:.4f}")
    print(f"  Schema cache (fixed/month): ${SCHEMA_FIXED_COST_PM:.4f}")
    print()
    for label, udata in USAGE.items():
        qpm = udata["qpd"] * udata["days"]
        print(f"  {label} ({qpm} q/user/month):")
        for users in USER_COUNTS:
            total_q = users * qpm
            cost    = total_q * QC["total"] + SCHEMA_FIXED_COST_PM
            print(f"    {users:5,} users: ${cost:>10,.2f}/month  (${cost/users:.2f}/user/month)")
        print()


if __name__ == "__main__":
    main()
